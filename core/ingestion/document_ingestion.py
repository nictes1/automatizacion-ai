#!/usr/bin/env python3
"""
Pipeline de ingesta de documentos
OCR/Visión → LLM Extractor → JSON normalizado → Validación → Guardado
"""

import os
import json
import logging
import requests
from typing import Dict, List, Any, Optional, Tuple
from pathlib import Path
import tempfile
import shutil

# Dependencias para procesamiento
try:
    import pytesseract
    from PIL import Image
    TESSERACT_AVAILABLE = True
except ImportError:
    TESSERACT_AVAILABLE = False

try:
    import pypdf
    PYPDF_AVAILABLE = True
except ImportError:
    PYPDF_AVAILABLE = False

try:
    from docx import Document
    DOCX_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False

# Configurar logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class DocumentIngestionPipeline:
    """Pipeline completo de ingesta de documentos"""
    
    def __init__(self, ollama_url: str = "http://localhost:11434", model: str = "llama3.1:8b"):
        self.ollama_url = ollama_url
        self.model = model
        self.supported_extensions = {
            # Texto/Tablas (alta prioridad)
            '.json', '.csv', '.tsv', '.xlsx', '.md', '.txt', '.html', '.docx', '.pdf',
            # Imágenes (condicional)
            '.png', '.jpg', '.jpeg', '.tiff', '.bmp',
            # Audio/Video (solo si se transcribe)
            '.mp3', '.m4a', '.wav', '.mp4', '.avi'
        }
    
    def is_supported_file(self, file_path: str) -> bool:
        """Verificar si el archivo es soportado"""
        ext = Path(file_path).suffix.lower()
        return ext in self.supported_extensions
    
    def get_file_priority(self, file_path: str) -> int:
        """Obtener prioridad del archivo (1=alta, 2=condicional, 3=baja)"""
        ext = Path(file_path).suffix.lower()
        
        # Alta prioridad
        if ext in ['.json', '.csv', '.tsv', '.xlsx', '.md', '.txt', '.html', '.docx']:
            return 1
        
        # Condicional (requiere OCR)
        if ext in ['.png', '.jpg', '.jpeg', '.tiff', '.bmp', '.pdf']:
            return 2
        
        # Baja prioridad (requiere transcripción)
        if ext in ['.mp3', '.m4a', '.wav', '.mp4', '.avi']:
            return 3
        
        return 0
    
    async def process_document(self, file_path: str, workspace_id: str, 
                             vertical: str) -> Dict[str, Any]:
        """Procesar documento completo"""
        try:
            logger.info(f"Procesando documento: {file_path}")
            
            # 1. Extraer texto
            text = await self._extract_text(file_path)
            if not text:
                return {"error": "No se pudo extraer texto del archivo"}
            
            # 2. Extraer estructura con LLM
            structured_data = await self._extract_structure(text, vertical)
            if not structured_data:
                return {"error": "No se pudo extraer estructura del documento"}
            
            # 3. Validar esquema
            is_valid = self._validate_schema(vertical, structured_data)
            if not is_valid:
                return {"error": "Estructura extraída no es válida"}
            
            # 4. Preparar para guardado
            result = {
                "success": True,
                "file_path": file_path,
                "workspace_id": workspace_id,
                "vertical": vertical,
                "structured_data": structured_data,
                "text_length": len(text),
                "extraction_method": self._get_extraction_method(file_path)
            }
            
            # 5. Separar texto libre para RAG
            rag_text = self._extract_rag_text(text, structured_data)
            if rag_text:
                result["rag_text"] = rag_text
            
            return result
            
        except Exception as e:
            logger.error(f"Error procesando documento: {e}")
            return {"error": str(e)}
    
    async def _extract_text(self, file_path: str) -> str:
        """Extraer texto del archivo"""
        ext = Path(file_path).suffix.lower()
        
        try:
            if ext == '.json':
                return self._extract_json_text(file_path)
            elif ext in ['.csv', '.tsv']:
                return self._extract_csv_text(file_path)
            elif ext == '.xlsx':
                return self._extract_xlsx_text(file_path)
            elif ext in ['.md', '.txt', '.html']:
                return self._extract_text_file(file_path)
            elif ext == '.docx':
                return self._extract_docx_text(file_path)
            elif ext == '.pdf':
                return self._extract_pdf_text(file_path)
            elif ext in ['.png', '.jpg', '.jpeg', '.tiff', '.bmp']:
                return self._extract_image_text(file_path)
            else:
                return ""
                
        except Exception as e:
            logger.error(f"Error extrayendo texto de {file_path}: {e}")
            return ""
    
    def _extract_json_text(self, file_path: str) -> str:
        """Extraer texto de JSON"""
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return json.dumps(data, indent=2, ensure_ascii=False)
    
    def _extract_csv_text(self, file_path: str) -> str:
        """Extraer texto de CSV/TSV"""
        import csv
        text = ""
        with open(file_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            for row in reader:
                text += " | ".join(row) + "\n"
        return text
    
    def _extract_xlsx_text(self, file_path: str) -> str:
        """Extraer texto de Excel"""
        import openpyxl
        text = ""
        workbook = openpyxl.load_workbook(file_path)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text += f"Hoja: {sheet_name}\n"
            for row in sheet.iter_rows():
                row_text = " | ".join([str(cell.value) for cell in row if cell.value])
                if row_text:
                    text += row_text + "\n"
        return text
    
    def _extract_text_file(self, file_path: str) -> str:
        """Extraer texto de archivo de texto"""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                return f.read()
        except UnicodeDecodeError:
            with open(file_path, 'r', encoding='latin-1') as f:
                return f.read()
    
    def _extract_docx_text(self, file_path: str) -> str:
        """Extraer texto de DOCX"""
        if not DOCX_AVAILABLE:
            return ""
        
        doc = Document(file_path)
        text = ""
        for paragraph in doc.paragraphs:
            text += paragraph.text + "\n"
        return text
    
    def _extract_pdf_text(self, file_path: str) -> str:
        """Extraer texto de PDF"""
        if not PYPDF_AVAILABLE:
            return ""
        
        text = ""
        with open(file_path, 'rb') as file:
            pdf_reader = pypdf.PdfReader(file)
            for page in pdf_reader.pages:
                text += page.extract_text() + "\n"
        return text
    
    def _extract_image_text(self, file_path: str) -> str:
        """Extraer texto de imagen usando OCR"""
        if not TESSERACT_AVAILABLE:
            return ""
        
        try:
            return pytesseract.image_to_string(Image.open(file_path))
        except Exception as e:
            logger.error(f"Error en OCR: {e}")
            return ""
    
    async def _extract_structure(self, text: str, vertical: str) -> Dict[str, Any]:
        """Extraer estructura usando LLM"""
        try:
            # Obtener esquema de ejemplo
            schema_example = self._get_schema_example(vertical)
            
            prompt = f"""
Analiza el siguiente texto y extrae la información estructurada según el esquema JSON del vertical "{vertical}".

TEXTO:
{text[:2000]}  # Limitar texto para evitar tokens excesivos

ESQUEMA ESPERADO:
{json.dumps(schema_example, indent=2, ensure_ascii=False)}

INSTRUCCIONES:
1. Extrae SOLO la información que esté presente en el texto
2. Mantén la estructura del esquema
3. Usa valores por defecto para campos faltantes
4. Valida que los precios sean números
5. Asegúrate de que los SKUs/IDs sean únicos
6. Responde SOLO con el JSON, sin texto adicional

JSON EXTRAÍDO:
"""
            
            response = requests.post(
                f"{self.ollama_url}/api/chat",
                json={
                    "model": self.model,
                    "messages": [
                        {"role": "user", "content": prompt}
                    ],
                    "stream": False
                },
                timeout=60
            )
            
            if response.status_code == 200:
                result = response.json()
                content = result.get("message", {}).get("content", "")
                
                # Extraer JSON de la respuesta
                json_start = content.find('{')
                json_end = content.rfind('}') + 1
                
                if json_start != -1 and json_end != -1:
                    json_str = content[json_start:json_end]
                    return json.loads(json_str)
                else:
                    logger.error("No se pudo extraer JSON del response del LLM")
                    return None
            else:
                logger.error(f"Error en LLM: {response.status_code}")
                return None
                
        except Exception as e:
            logger.error(f"Error extrayendo estructura: {e}")
            return None
    
    def _get_schema_example(self, vertical: str) -> Dict[str, Any]:
        """Obtener ejemplo de esquema por vertical"""
        from core.schemas.vertical_schemas import get_schema_example
        return get_schema_example(vertical)
    
    def _validate_schema(self, vertical: str, data: Dict[str, Any]) -> bool:
        """Validar esquema extraído"""
        from core.schemas.vertical_schemas import validate_schema
        return validate_schema(vertical, data)
    
    def _get_extraction_method(self, file_path: str) -> str:
        """Obtener método de extracción usado"""
        ext = Path(file_path).suffix.lower()
        
        if ext in ['.json', '.csv', '.tsv', '.xlsx']:
            return "structured"
        elif ext in ['.md', '.txt', '.html', '.docx']:
            return "text"
        elif ext == '.pdf':
            return "pdf_text"
        elif ext in ['.png', '.jpg', '.jpeg', '.tiff', '.bmp']:
            return "ocr"
        else:
            return "unknown"
    
    def _extract_rag_text(self, original_text: str, structured_data: Dict[str, Any]) -> str:
        """Extraer texto libre para RAG (políticas, FAQ, etc.)"""
        # En una implementación real, esto separaría el texto que no se pudo estructurar
        # y que es útil para RAG (políticas, FAQ, descripciones largas)
        
        rag_sections = []
        
        # Buscar secciones que no están en la estructura
        if "policies" in original_text.lower():
            rag_sections.append("Políticas y condiciones")
        
        if "faq" in original_text.lower() or "preguntas" in original_text.lower():
            rag_sections.append("Preguntas frecuentes")
        
        if "descripción" in original_text.lower() or "descripcion" in original_text.lower():
            rag_sections.append("Descripciones detalladas")
        
        return "\n\n".join(rag_sections) if rag_sections else ""

class DocumentProcessor:
    """Procesador de documentos con chunking para RAG"""
    
    def __init__(self, chunk_size: int = 400, overlap: int = 50):
        self.chunk_size = chunk_size
        self.overlap = overlap
    
    def chunk_text(self, text: str, metadata: Dict[str, Any] = None) -> List[Dict[str, Any]]:
        """Dividir texto en chunks para RAG"""
        if not text.strip():
            return []
        
        # Chunking simple por párrafos
        paragraphs = text.split('\n\n')
        chunks = []
        
        current_chunk = ""
        chunk_index = 0
        
        for paragraph in paragraphs:
            if len(current_chunk) + len(paragraph) <= self.chunk_size:
                current_chunk += paragraph + "\n\n"
            else:
                if current_chunk:
                    chunks.append(self._create_chunk(current_chunk, chunk_index, metadata))
                    chunk_index += 1
                
                # Si el párrafo es muy largo, dividirlo
                if len(paragraph) > self.chunk_size:
                    sub_chunks = self._split_long_paragraph(paragraph, metadata, chunk_index)
                    chunks.extend(sub_chunks)
                    chunk_index += len(sub_chunks)
                    current_chunk = ""
                else:
                    current_chunk = paragraph + "\n\n"
        
        # Agregar último chunk
        if current_chunk:
            chunks.append(self._create_chunk(current_chunk, chunk_index, metadata))
        
        return chunks
    
    def _create_chunk(self, content: str, index: int, metadata: Dict[str, Any] = None) -> Dict[str, Any]:
        """Crear chunk con metadatos"""
        chunk_metadata = {
            "chunk_index": index,
            "type": "rag_content"
        }
        
        if metadata:
            chunk_metadata.update(metadata)
        
        return {
            "content": content.strip(),
            "metadata": chunk_metadata
        }
    
    def _split_long_paragraph(self, paragraph: str, metadata: Dict[str, Any], start_index: int) -> List[Dict[str, Any]]:
        """Dividir párrafo largo en chunks más pequeños"""
        words = paragraph.split()
        chunks = []
        current_chunk = ""
        chunk_index = start_index
        
        for word in words:
            if len(current_chunk) + len(word) + 1 <= self.chunk_size:
                current_chunk += word + " "
            else:
                if current_chunk:
                    chunks.append(self._create_chunk(current_chunk, chunk_index, metadata))
                    chunk_index += 1
                current_chunk = word + " "
        
        if current_chunk:
            chunks.append(self._create_chunk(current_chunk, chunk_index, metadata))
        
        return chunks

# Función de conveniencia
def create_ingestion_pipeline(ollama_url: str = "http://localhost:11434", 
                            model: str = "llama3.1:8b") -> DocumentIngestionPipeline:
    """Crear pipeline de ingesta"""
    return DocumentIngestionPipeline(ollama_url, model)

def create_document_processor(chunk_size: int = 400, overlap: int = 50) -> DocumentProcessor:
    """Crear procesador de documentos"""
    return DocumentProcessor(chunk_size, overlap)

